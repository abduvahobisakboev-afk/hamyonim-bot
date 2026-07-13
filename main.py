"""
==============================================================================
 HAMYONIM — Karta-markazlashgan Shaxsiy Moliya Boti (BITTA FAYLLI VERSIYA)
==============================================================================
Bu fayl butun botni (konfiguratsiya, ma'lumotlar bazasi, klaviaturalar,
SMS-webhook, va barcha handlerlar) BITTA main.py faylida jamlaydi.

Ishga tushirish:
    pip install -r requirements.txt
    export HAMYONIM_ADMIN_IDS="sizning_telegram_id_raqamingiz"
    python3 main.py
==============================================================================
"""

from __future__ import annotations

import asyncio
import io
import re
import secrets
import sqlite3
import sys
from contextlib import closing
from datetime import datetime, timedelta
from enum import Enum
from typing import Any, Optional

from aiohttp import web
from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramAPIError, TelegramForbiddenError, TelegramRetryAfter
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BotCommand, BufferedInputFile, CallbackQuery, InlineKeyboardButton,
    InlineKeyboardMarkup, KeyboardButton, Message, ReplyKeyboardMarkup,
)
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
import logging
import os


# ======================================================================
# 1. KONFIGURATSIYA
# ======================================================================


# ------------------------------------------------------------------ BOT
# Diqqat: tokenni ochiq kodga yozish xavfsiz emas (kod boshqa birov qo'liga
# o'tsa, botingizni egallab olishi mumkin). Shu sabab avval muhit
# o'zgaruvchisidan o'qiladi, topilmasa quyidagi standart qiymat ishlatiladi.
BOT_TOKEN: str = os.getenv("HAMYONIM_BOT_TOKEN", "8888847127:AAEbgW0Kk97WPRyqdZaslynlxrNrbE1vNa0")

# Bot egasi (admin) — barcha kartalar va statistikaga to'liq huquqga ega
ADMIN_IDS: set[int] = {
    int(uid) for uid in os.getenv("HAMYONIM_ADMIN_IDS", "1673990832").split(",") if uid.strip().isdigit()
}

# ------------------------------------------------------------------ DB / LOG
DB_PATH: str = os.getenv("HAMYONIM_DB_PATH", "hamyonim.db")
LOG_PATH: str = os.getenv("HAMYONIM_LOG_PATH", "hamyonim.log")

# ------------------------------------------------------------------ WEB SERVER (Render port + SMS webhook)
PORT: int = int(os.getenv("PORT", "10000"))

# SMS-webhook uchun maxfiy kalit — faqat shu kalitni biladigan so'rovlar qabul qilinadi.
# Muhim: buni SMS-forward ilovangizga ham kiritishingiz kerak bo'ladi.
SMS_WEBHOOK_SECRET: str = os.getenv("HAMYONIM_SMS_SECRET", secrets.token_urlsafe(16))

TRANSACTIONS_PER_PAGE: int = 8
BROADCAST_DELAY_SECONDS: float = 0.05
USERS_PER_BROADCAST_BATCH: int = 25

# ------------------------------------------------------------------ BYUDJET
BUDGET_WARNING_THRESHOLD: float = 0.8   # 80% sarflanganda ogohlantirish
BUDGET_EXCEEDED_THRESHOLD: float = 1.0  # 100% dan oshganda qattiq ogohlantirish


# ======================================================================
# 2. LOGGING
# ======================================================================


def setup_logging() -> logging.Logger:
    logger = logging.getLogger("hamyonim")
    logger.setLevel(logging.INFO)

    if logger.handlers:
        return logger

    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    console_handler.setLevel(logging.INFO)

    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.INFO)

    logger.addHandler(console_handler)
    logger.addHandler(file_handler)
    return logger


logger = setup_logging()


# ======================================================================
# 3. YORDAMCHI FUNKSIYALAR
# ======================================================================
class CardType(str, Enum):
    UZCARD = "Uzcard"
    HUMO = "Humo"
    UNKNOWN = "Noma'lum"


def clean_card_number(raw: str) -> str:
    return re.sub(r"[^\d]", "", raw)


def detect_card_type(card_number: str) -> CardType:
    digits = clean_card_number(card_number)
    if digits.startswith("8600"):
        return CardType.UZCARD
    if digits.startswith("9860"):
        return CardType.HUMO
    return CardType.UNKNOWN


def is_valid_card_number(card_number: str) -> bool:
    digits = clean_card_number(card_number)
    return len(digits) == 16 and digits.isdigit()


def mask_card_number(card_number: str) -> str:
    digits = clean_card_number(card_number)
    if len(digits) != 16:
        return "**** **** **** ****"
    return f"{digits[:4]} **** **** {digits[12:]}"


def last4(card_number: str) -> str:
    digits = clean_card_number(card_number)
    return digits[-4:] if len(digits) >= 4 else digits


def format_money(amount: float) -> str:
    sign = "-" if amount < 0 else ""
    amount = abs(amount)
    integer_part = int(amount)
    decimal_part = round((amount - integer_part) * 100)
    grouped = f"{integer_part:,}".replace(",", " ")
    return f"{sign}{grouped}.{decimal_part:02d} so'm"


def format_datetime(dt_str: str) -> str:
    try:
        dt = datetime.fromisoformat(dt_str)
        return dt.strftime("%d.%m.%Y %H:%M")
    except (ValueError, TypeError):
        return dt_str


def escape_html(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def safe_float(raw: str) -> Optional[float]:
    cleaned = raw.strip().replace(" ", "").replace(",", ".")
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


# ======================================================================
# 4. MA'LUMOTLAR BAZASI
# ======================================================================
class Database:
    def __init__(self, db_path: str = DB_PATH) -> None:
        self.db_path = db_path
        self._init_db()

    def _get_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA foreign_keys=ON;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        return conn

    def _init_db(self) -> None:
        with closing(self._get_connection()) as conn, conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS users (
                    user_id     INTEGER PRIMARY KEY,
                    username    TEXT,
                    full_name   TEXT,
                    is_blocked  INTEGER NOT NULL DEFAULT 0,
                    created_at  TEXT NOT NULL,
                    updated_at  TEXT NOT NULL
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS cards (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    owner_user_id   INTEGER,
                    card_number     TEXT NOT NULL UNIQUE,
                    card_type       TEXT NOT NULL,
                    balance         REAL NOT NULL DEFAULT 0,
                    added_by_admin  INTEGER,
                    created_at      TEXT NOT NULL
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS transactions (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    card_id     INTEGER NOT NULL,
                    type        TEXT NOT NULL CHECK (type IN ('income', 'expense')),
                    amount      REAL NOT NULL,
                    counterparty TEXT,
                    raw_sms     TEXT,
                    created_at  TEXT NOT NULL,
                    FOREIGN KEY (card_id) REFERENCES cards (id) ON DELETE CASCADE
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS card_budgets (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    card_id       INTEGER NOT NULL UNIQUE,
                    monthly_limit REAL NOT NULL,
                    created_at    TEXT NOT NULL,
                    FOREIGN KEY (card_id) REFERENCES cards (id) ON DELETE CASCADE
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS savings_goals (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id         INTEGER NOT NULL,
                    name            TEXT NOT NULL,
                    target_amount   REAL NOT NULL,
                    current_amount  REAL NOT NULL DEFAULT 0,
                    deadline        TEXT,
                    completed       INTEGER NOT NULL DEFAULT 0,
                    created_at      TEXT NOT NULL
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS contact_relay (
                    admin_message_id  INTEGER NOT NULL,
                    admin_chat_id     INTEGER NOT NULL,
                    source_user_id    INTEGER NOT NULL,
                    created_at        TEXT NOT NULL,
                    PRIMARY KEY (admin_message_id, admin_chat_id)
                );
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_cards_owner ON cards (owner_user_id);")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_tx_card_id ON transactions (card_id);")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_tx_created_at ON transactions (created_at);")
        logger.info("Ma'lumotlar bazasi muvaffaqiyatli ishga tushirildi: %s", self.db_path)

    # ================================================================== USERS

    def upsert_user(self, user_id: int, username: Optional[str], full_name: str) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn, conn:
            conn.execute(
                """
                INSERT INTO users (user_id, username, full_name, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(user_id) DO UPDATE SET
                    username = excluded.username,
                    full_name = excluded.full_name,
                    updated_at = excluded.updated_at;
                """,
                (user_id, username, full_name, now, now),
            )

    def get_user(self, user_id: int) -> Optional[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute("SELECT * FROM users WHERE user_id = ?;", (user_id,)).fetchone()

    def user_exists(self, user_id: int) -> bool:
        return self.get_user(user_id) is not None

    def set_blocked(self, user_id: int, blocked: bool) -> None:
        with closing(self._get_connection()) as conn, conn:
            conn.execute("UPDATE users SET is_blocked = ? WHERE user_id = ?;", (1 if blocked else 0, user_id))

    def get_all_user_ids(self, only_active: bool = True) -> list[int]:
        query = "SELECT user_id FROM users"
        if only_active:
            query += " WHERE is_blocked = 0"
        with closing(self._get_connection()) as conn:
            return [row["user_id"] for row in conn.execute(query + ";").fetchall()]

    def count_users(self) -> int:
        with closing(self._get_connection()) as conn:
            return conn.execute("SELECT COUNT(*) AS cnt FROM users;").fetchone()["cnt"]

    def count_active_users_today(self) -> int:
        today = datetime.now().strftime("%Y-%m-%d")
        with closing(self._get_connection()) as conn:
            return conn.execute(
                "SELECT COUNT(*) AS cnt FROM users WHERE updated_at LIKE ?;", (f"{today}%",)
            ).fetchone()["cnt"]

    # ================================================================== CARDS

    def add_card(self, owner_user_id: Optional[int], card_number: str, added_by_admin: Optional[int]) -> int:
        card_type = detect_card_type(card_number).value
        now = datetime.now().isoformat(timespec="seconds")
        try:
            with closing(self._get_connection()) as conn, conn:
                cur = conn.execute(
                    """
                    INSERT INTO cards (owner_user_id, card_number, card_type, balance, added_by_admin, created_at)
                    VALUES (?, ?, ?, 0, ?, ?);
                    """,
                    (owner_user_id, clean_card_number(card_number), card_type, added_by_admin, now),
                )
                return cur.lastrowid
        except sqlite3.IntegrityError as e:
            raise ValueError("CARD_ALREADY_EXISTS") from e

    def get_cards_by_owner(self, owner_user_id: int) -> list[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute(
                "SELECT * FROM cards WHERE owner_user_id = ? ORDER BY created_at DESC;", (owner_user_id,)
            ).fetchall()

    def find_card_by_full_number(self, card_number: str) -> Optional[sqlite3.Row]:
        """Karta to'liq raqami bo'yicha (ID emas) qidiradi — egasini avtomatik aniqlash uchun."""
        cleaned = clean_card_number(card_number)
        with closing(self._get_connection()) as conn:
            return conn.execute("SELECT * FROM cards WHERE card_number = ?;", (cleaned,)).fetchone()

    def assign_card_owner(self, card_id: int, owner_user_id: int) -> None:
        """Admin oldindan (egasiz) qo'shgan kartani, haqiqiy egasi raqamni kiritganda unga bog'laydi."""
        with closing(self._get_connection()) as conn, conn:
            conn.execute("UPDATE cards SET owner_user_id = ? WHERE id = ?;", (owner_user_id, card_id))

    def get_card(self, card_id: int) -> Optional[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute("SELECT * FROM cards WHERE id = ?;", (card_id,)).fetchone()

    def get_card_owned_by(self, card_id: int, owner_user_id: int) -> Optional[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute(
                "SELECT * FROM cards WHERE id = ? AND owner_user_id = ?;", (card_id, owner_user_id)
            ).fetchone()

    def get_all_cards(self) -> list[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute("SELECT * FROM cards ORDER BY created_at DESC;").fetchall()

    def find_cards_by_last4(self, last4_digits: str) -> list[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute(
                "SELECT * FROM cards WHERE substr(card_number, -4) = ?;", (last4_digits,)
            ).fetchall()

    def count_cards(self, owner_user_id: Optional[int] = None) -> int:
        with closing(self._get_connection()) as conn:
            if owner_user_id is None:
                return conn.execute("SELECT COUNT(*) AS cnt FROM cards;").fetchone()["cnt"]
            return conn.execute(
                "SELECT COUNT(*) AS cnt FROM cards WHERE owner_user_id = ?;", (owner_user_id,)
            ).fetchone()["cnt"]

    def delete_card(self, card_id: int) -> bool:
        with closing(self._get_connection()) as conn, conn:
            cur = conn.execute("DELETE FROM cards WHERE id = ?;", (card_id,))
            return cur.rowcount > 0

    def update_card_balance(self, card_id: int, delta: float) -> float:
        with closing(self._get_connection()) as conn, conn:
            conn.execute("UPDATE cards SET balance = balance + ? WHERE id = ?;", (delta, card_id))
            row = conn.execute("SELECT balance FROM cards WHERE id = ?;", (card_id,)).fetchone()
            return row["balance"]

    def get_total_balance(self, owner_user_id: int) -> float:
        with closing(self._get_connection()) as conn:
            row = conn.execute(
                "SELECT COALESCE(SUM(balance), 0) AS total FROM cards WHERE owner_user_id = ?;", (owner_user_id,)
            ).fetchone()
            return row["total"]

    # ============================================================ TRANSACTIONS

    def add_transaction(
        self, card_id: int, tx_type: str, amount: float,
        counterparty: Optional[str] = None, raw_sms: Optional[str] = None,
    ) -> int:
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn, conn:
            cur = conn.execute(
                """
                INSERT INTO transactions (card_id, type, amount, counterparty, raw_sms, created_at)
                VALUES (?, ?, ?, ?, ?, ?);
                """,
                (card_id, tx_type, amount, counterparty, raw_sms, now),
            )
            tx_id = cur.lastrowid
        self.update_card_balance(card_id, amount if tx_type == "income" else -amount)
        return tx_id

    def get_transactions_by_card(self, card_id: int, limit: int = 8, offset: int = 0) -> list[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute(
                """
                SELECT * FROM transactions WHERE card_id = ?
                ORDER BY created_at DESC, id DESC LIMIT ? OFFSET ?;
                """,
                (card_id, limit, offset),
            ).fetchall()

    def count_transactions_by_card(self, card_id: int) -> int:
        with closing(self._get_connection()) as conn:
            return conn.execute(
                "SELECT COUNT(*) AS cnt FROM transactions WHERE card_id = ?;", (card_id,)
            ).fetchone()["cnt"]

    def get_all_transactions_for_user(self, user_id: int) -> list[sqlite3.Row]:
        """Foydalanuvchining BARCHA kartalaridagi tranzaksiyalarini (eksport uchun) qaytaradi."""
        with closing(self._get_connection()) as conn:
            return conn.execute(
                """
                SELECT t.*, c.card_type, c.card_number FROM transactions t
                JOIN cards c ON c.id = t.card_id
                WHERE c.owner_user_id = ?
                ORDER BY t.created_at ASC;
                """,
                (user_id,),
            ).fetchall()

    def get_stats_summary_for_user(self, user_id: int, days: int) -> dict[str, float]:
        since = (datetime.now() - timedelta(days=days)).isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn:
            rows = conn.execute(
                """
                SELECT t.type, COALESCE(SUM(t.amount), 0) AS total FROM transactions t
                JOIN cards c ON c.id = t.card_id
                WHERE c.owner_user_id = ? AND t.created_at >= ?
                GROUP BY t.type;
                """,
                (user_id, since),
            ).fetchall()
        result = {"income": 0.0, "expense": 0.0}
        for row in rows:
            result[row["type"]] = row["total"]
        return result

    def get_counterparty_breakdown_for_user(self, user_id: int, tx_type: str, days: int) -> list[sqlite3.Row]:
        since = (datetime.now() - timedelta(days=days)).isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn:
            return conn.execute(
                """
                SELECT COALESCE(t.counterparty, 'Noma\'lum') AS counterparty,
                       COALESCE(SUM(t.amount), 0) AS total, COUNT(*) AS cnt
                FROM transactions t
                JOIN cards c ON c.id = t.card_id
                WHERE c.owner_user_id = ? AND t.type = ? AND t.created_at >= ?
                GROUP BY counterparty ORDER BY total DESC;
                """,
                (user_id, tx_type, since),
            ).fetchall()

    # =========================================================== ADMIN STATS

    def get_global_stats(self) -> dict[str, Any]:
        with closing(self._get_connection()) as conn:
            total_users = conn.execute("SELECT COUNT(*) AS cnt FROM users;").fetchone()["cnt"]
            total_cards = conn.execute("SELECT COUNT(*) AS cnt FROM cards;").fetchone()["cnt"]
            total_tx = conn.execute("SELECT COUNT(*) AS cnt FROM transactions;").fetchone()["cnt"]
            total_income = conn.execute(
                "SELECT COALESCE(SUM(amount), 0) AS s FROM transactions WHERE type='income';"
            ).fetchone()["s"]
            total_expense = conn.execute(
                "SELECT COALESCE(SUM(amount), 0) AS s FROM transactions WHERE type='expense';"
            ).fetchone()["s"]
            total_balance = conn.execute("SELECT COALESCE(SUM(balance), 0) AS s FROM cards;").fetchone()["s"]
            blocked = conn.execute("SELECT COUNT(*) AS cnt FROM users WHERE is_blocked = 1;").fetchone()["cnt"]
        return {
            "total_users": total_users,
            "blocked_users": blocked,
            "total_cards": total_cards,
            "total_transactions": total_tx,
            "total_income": total_income,
            "total_expense": total_expense,
            "total_balance": total_balance,
        }


    # ================================================================ BUDGETS

    def set_card_budget(self, card_id: int, monthly_limit: float) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn, conn:
            conn.execute(
                """
                INSERT INTO card_budgets (card_id, monthly_limit, created_at) VALUES (?, ?, ?)
                ON CONFLICT(card_id) DO UPDATE SET monthly_limit = excluded.monthly_limit;
                """,
                (card_id, monthly_limit, now),
            )

    def get_card_budget(self, card_id: int) -> Optional[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute("SELECT * FROM card_budgets WHERE card_id = ?;", (card_id,)).fetchone()

    def delete_card_budget(self, card_id: int) -> bool:
        with closing(self._get_connection()) as conn, conn:
            cur = conn.execute("DELETE FROM card_budgets WHERE card_id = ?;", (card_id,))
            return cur.rowcount > 0

    def get_card_spent_this_month(self, card_id: int) -> float:
        month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0).isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn:
            row = conn.execute(
                """
                SELECT COALESCE(SUM(amount), 0) AS total FROM transactions
                WHERE card_id = ? AND type = 'expense' AND created_at >= ?;
                """,
                (card_id, month_start),
            ).fetchone()
            return row["total"]

    # ======================================================= SAVINGS GOALS

    def add_goal(self, user_id: int, name: str, target_amount: float, deadline: Optional[str]) -> int:
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn, conn:
            cur = conn.execute(
                """
                INSERT INTO savings_goals (user_id, name, target_amount, current_amount, deadline, created_at)
                VALUES (?, ?, ?, 0, ?, ?);
                """,
                (user_id, name, target_amount, deadline, now),
            )
            return cur.lastrowid

    def get_goals(self, user_id: int) -> list[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute(
                "SELECT * FROM savings_goals WHERE user_id = ? ORDER BY created_at DESC;", (user_id,)
            ).fetchall()

    def get_goal(self, goal_id: int, user_id: int) -> Optional[sqlite3.Row]:
        with closing(self._get_connection()) as conn:
            return conn.execute(
                "SELECT * FROM savings_goals WHERE id = ? AND user_id = ?;", (goal_id, user_id)
            ).fetchone()

    def add_to_goal(self, goal_id: int, user_id: int, amount: float) -> Optional[sqlite3.Row]:
        goal = self.get_goal(goal_id, user_id)
        if goal is None:
            return None
        new_amount = goal["current_amount"] + amount
        completed = 1 if new_amount >= goal["target_amount"] else 0
        with closing(self._get_connection()) as conn, conn:
            conn.execute(
                "UPDATE savings_goals SET current_amount = ?, completed = ? WHERE id = ?;",
                (new_amount, completed, goal_id),
            )
        return self.get_goal(goal_id, user_id)

    def delete_goal(self, goal_id: int, user_id: int) -> bool:
        with closing(self._get_connection()) as conn, conn:
            cur = conn.execute("DELETE FROM savings_goals WHERE id = ? AND user_id = ?;", (goal_id, user_id))
            return cur.rowcount > 0

    # ========================================================= CONTACT RELAY

    def save_contact_relay(self, admin_message_id: int, admin_chat_id: int, source_user_id: int) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        with closing(self._get_connection()) as conn, conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO contact_relay (admin_message_id, admin_chat_id, source_user_id, created_at)
                VALUES (?, ?, ?, ?);
                """,
                (admin_message_id, admin_chat_id, source_user_id, now),
            )

    def get_contact_relay_source(self, admin_message_id: int, admin_chat_id: int) -> Optional[int]:
        with closing(self._get_connection()) as conn:
            row = conn.execute(
                "SELECT source_user_id FROM contact_relay WHERE admin_message_id = ? AND admin_chat_id = ?;",
                (admin_message_id, admin_chat_id),
            ).fetchone()
            return row["source_user_id"] if row else None


db = Database(DB_PATH)


# ======================================================================
# 5. FSM HOLATLARI
# ======================================================================
class AdminAddCardStates(StatesGroup):
    entering_card_number = State()  # Admin faqat karta raqamini kiritadi — egasi keyin avtomatik aniqlanadi


class SelfAddCardStates(StatesGroup):
    entering_card_number = State()  # Har qanday foydalanuvchi o'z kartasini o'zi qo'shadi


class ContactAdminStates(StatesGroup):
    entering_message = State()  # Foydalanuvchi adminga yubormoqchi bo'lgan xabarni yozadi


class BroadcastStates(StatesGroup):
    entering_message = State()
    confirming = State()


class BudgetStates(StatesGroup):
    choosing_card = State()
    entering_limit = State()


class GoalStates(StatesGroup):
    entering_name = State()
    entering_target = State()
    entering_deadline = State()


class GoalDepositStates(StatesGroup):
    entering_amount = State()


# ======================================================================
# 6. KLAVIATURALAR
# ======================================================================
def main_menu_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    builder = ReplyKeyboardBuilder()
    builder.add(KeyboardButton(text="🔄 Start"))
    builder.add(KeyboardButton(text="💰 Balans"))
    builder.add(KeyboardButton(text="🧾 Tarix"))
    builder.add(KeyboardButton(text="💳 Kartalarim"))
    builder.add(KeyboardButton(text="➕ Karta qo'shish"))
    builder.add(KeyboardButton(text="📊 Statistika"))
    builder.add(KeyboardButton(text="🎯 Byudjet"))
    builder.add(KeyboardButton(text="🏆 Maqsadlar"))
    builder.add(KeyboardButton(text="📁 Eksport"))
    builder.adjust(2, 2, 2, 2, 1)

    if user_id in ADMIN_IDS:
        builder.row(KeyboardButton(text="🛠 Admin Panel"))
    else:
        builder.row(KeyboardButton(text="✉️ Admin bilan bog'lanish"))

    builder.row(KeyboardButton(text="🚫 Bekor qilish"))
    return builder.as_markup(resize_keyboard=True, input_field_placeholder="Menyudan tanlang...")


def cancel_keyboard() -> ReplyKeyboardMarkup:
    builder = ReplyKeyboardBuilder()
    builder.add(KeyboardButton(text="🚫 Bekor qilish"))
    return builder.as_markup(resize_keyboard=True)


# --------------------------------------------------------------- KARTALAR

def cards_list_keyboard(cards: Iterable[sqlite3.Row], for_admin: bool = False) -> InlineKeyboardMarkup:
    """
    Kartalar ro'yxatini tugmalar sifatida ko'rsatadi. Har bir tugmada
    bank turi, niqoblangan raqam va joriy balans ko'rinadi.
    """
    builder = InlineKeyboardBuilder()
    for card in cards:
        text = f"{'💳' if card['card_type']=='Uzcard' else '🟣'} {card['card_type']} {mask_card_number(card['card_number'])} | {format_money(card['balance'])}"
        builder.row(InlineKeyboardButton(text=text, callback_data=f"card_hist:{card['id']}:0"))

    builder.row(InlineKeyboardButton(text="➕ Yangi karta qo'shish", callback_data="self_add_card"))

    if for_admin:
        for card in cards:
            builder.row(
                InlineKeyboardButton(
                    text=f"🗑 {mask_card_number(card['card_number'])} o'chirish", callback_data=f"admin_card_del:{card['id']}"
                )
            )

    return builder.as_markup()


def card_history_pagination_keyboard(card_id: int, current_page: int, total_pages: int) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    nav_row: list[InlineKeyboardButton] = []
    if current_page > 0:
        nav_row.append(InlineKeyboardButton(text="⬅️", callback_data=f"card_hist:{card_id}:{current_page - 1}"))
    nav_row.append(InlineKeyboardButton(text=f"{current_page + 1}/{max(total_pages, 1)}", callback_data="tx_noop"))
    if current_page < total_pages - 1:
        nav_row.append(InlineKeyboardButton(text="➡️", callback_data=f"card_hist:{card_id}:{current_page + 1}"))
    builder.row(*nav_row)
    builder.row(InlineKeyboardButton(text="⬅️ Kartalarga qaytish", callback_data="cards_back"))
    return builder.as_markup()


def confirm_yes_no_keyboard(yes_cb: str, no_cb: str) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="✅ Ha", callback_data=yes_cb))
    builder.add(InlineKeyboardButton(text="❌ Yo'q", callback_data=no_cb))
    builder.adjust(2)
    return builder.as_markup()


# ----------------------------------------------------------------- ADMIN

def admin_panel_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="➕ Karta qo'shish", callback_data="admin_add_card"))
    builder.row(InlineKeyboardButton(text="💳 Barcha kartalar", callback_data="admin_all_cards"))
    builder.row(InlineKeyboardButton(text="📊 Statistika", callback_data="admin_stats"))
    builder.row(InlineKeyboardButton(text="📢 Xabar tarqatish", callback_data="admin_broadcast"))
    return builder.as_markup()


def broadcast_confirm_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="✅ Yuborish", callback_data="broadcast_confirm"))
    builder.add(InlineKeyboardButton(text="❌ Bekor qilish", callback_data="broadcast_cancel"))
    builder.adjust(2)
    return builder.as_markup()


def add_another_card_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="➕ Yana karta qo'shish", callback_data="admin_add_card"))
    builder.add(InlineKeyboardButton(text="✅ Tugatish", callback_data="admin_add_card_done"))
    builder.adjust(1)
    return builder.as_markup()


# --------------------------------------------------------------- BYUDJET

def budget_card_choice_keyboard(cards: Iterable[sqlite3.Row]) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    for c in cards:
        builder.row(
            InlineKeyboardButton(
                text=f"{c['card_type']} {mask_card_number(c['card_number'])}", callback_data=f"budget_card:{c['id']}"
            )
        )
    builder.row(InlineKeyboardButton(text="🚫 Bekor qilish", callback_data="budget_cancel"))
    return builder.as_markup()


def budget_list_keyboard(items: list[tuple]) -> InlineKeyboardMarkup:
    """items: [(card_row, budget_row_or_None), ...]"""
    builder = InlineKeyboardBuilder()
    for card, budget in items:
        if budget is not None:
            builder.row(
                InlineKeyboardButton(
                    text=f"🗑 {card['card_type']} {mask_card_number(card['card_number'])}",
                    callback_data=f"budget_del:{card['id']}",
                )
            )
    builder.row(InlineKeyboardButton(text="➕ Yangi byudjet", callback_data="budget_add"))
    return builder.as_markup()


# ---------------------------------------------------------------- GOALS

def goals_list_keyboard(goals: Iterable[sqlite3.Row]) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    for g in goals:
        status = "✅" if g["completed"] else "🏆"
        builder.row(InlineKeyboardButton(text=f"{status} {g['name']}", callback_data=f"goal_view:{g['id']}"))
    builder.row(InlineKeyboardButton(text="➕ Yangi maqsad", callback_data="goal_add"))
    return builder.as_markup()


def goal_detail_keyboard(goal_id: int) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="💵 Mablag' qo'shish", callback_data=f"goal_deposit:{goal_id}"))
    builder.add(InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"goal_del:{goal_id}"))
    builder.adjust(1)
    builder.row(InlineKeyboardButton(text="⬅️ Orqaga", callback_data="goal_back"))
    return builder.as_markup()


# --------------------------------------------------------------- STATISTIKA

def stats_period_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="7 kun", callback_data="stats_period:7"))
    builder.add(InlineKeyboardButton(text="30 kun", callback_data="stats_period:30"))
    builder.add(InlineKeyboardButton(text="90 kun", callback_data="stats_period:90"))
    builder.adjust(3)
    builder.row(InlineKeyboardButton(text="📊 Grafik", callback_data="stats_chart"))
    return builder.as_markup()


# --------------------------------------------------------------- EKSPORT

def export_format_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="📊 Excel (.xlsx)", callback_data="export_xlsx"))
    builder.add(InlineKeyboardButton(text="📄 PDF", callback_data="export_pdf"))
    builder.adjust(2)
    return builder.as_markup()


# ======================================================================
# 7. SMS-WEBHOOK VA PARSER
# ======================================================================
# ------------------------------------------------------------------ PARSER

INCOME_KEYWORDS = re.compile(
    r"hisobi?ngizga|qo['`]?shildi|zachislen|popolnen|kredit|kirim|to'?ldirildi",
    re.IGNORECASE,
)
EXPENSE_KEYWORDS = re.compile(
    r"hisobi?ngizdan|yechib\s*olindi|spisan|to'?lov|xarid|chiqim|debet|yechildi",
    re.IGNORECASE,
)
AMOUNT_PATTERN = re.compile(r"(\d[\d\s.,]{2,})\s*(so['`]?m|uzs)", re.IGNORECASE)
CARD_LAST4_PATTERN = re.compile(r"\*\s?(\d{4})\b")
CARD_LAST4_FALLBACK = re.compile(r"karta\D{0,15}(\d{4})\b", re.IGNORECASE)
COUNTERPARTY_PATTERN = re.compile(
    r"(?:do['`]?kon|merchant|qabul\s*qiluvchi|kimga|oluvchi)\s*[:\-]?\s*([^\n,.;]{2,40})",
    re.IGNORECASE,
)


class ParsedSms:
    def __init__(self, tx_type: str, amount: float, card_last4: str, counterparty: Optional[str]) -> None:
        self.tx_type = tx_type
        self.amount = amount
        self.card_last4 = card_last4
        self.counterparty = counterparty


def parse_bank_sms(text: str) -> Optional[ParsedSms]:
    """SMS matnidan tur, summa, karta oxirgi 4 raqami va kontragentni ajratib oladi."""
    is_income = bool(INCOME_KEYWORDS.search(text))
    is_expense = bool(EXPENSE_KEYWORDS.search(text))

    if is_income and not is_expense:
        tx_type = "income"
    elif is_expense and not is_income:
        tx_type = "expense"
    else:
        return None  # Turi noaniq — xavfsizlik uchun qayta ishlanmaydi

    amount_match = AMOUNT_PATTERN.search(text)
    if not amount_match:
        return None
    amount_str = re.sub(r"[^\d.]", "", amount_match.group(1).replace(",", "."))
    try:
        amount = float(amount_str)
    except ValueError:
        return None
    if amount <= 0:
        return None

    card_match = CARD_LAST4_PATTERN.search(text) or CARD_LAST4_FALLBACK.search(text)
    if not card_match:
        return None
    card_last4 = card_match.group(1)

    cp_match = COUNTERPARTY_PATTERN.search(text)
    counterparty = cp_match.group(1).strip() if cp_match else None

    return ParsedSms(tx_type=tx_type, amount=amount, card_last4=card_last4, counterparty=counterparty)


# --------------------------------------------------------------- HTTP APP

def create_web_app(bot: Bot) -> web.Application:
    app = web.Application()

    async def health(request: web.Request) -> web.Response:
        return web.Response(text="Hamyonim bot ishlayapti ✅")

    async def sms_webhook(request: web.Request) -> web.Response:
        try:
            payload = await request.json()
        except Exception:
            return web.json_response({"ok": False, "error": "JSON kutilgan edi"}, status=400)

        if payload.get("secret") != SMS_WEBHOOK_SECRET:
            logger.warning("SMS webhookka noto'g'ri secret bilan urinish qilindi.")
            return web.json_response({"ok": False, "error": "Noto'g'ri secret"}, status=403)

        text = (payload.get("text") or "").strip()
        if not text:
            return web.json_response({"ok": False, "error": "text bo'sh"}, status=400)

        parsed = parse_bank_sms(text)
        if parsed is None:
            logger.info("SMS tanib bo'lmadi: %s", text[:200])
            return web.json_response({"ok": False, "error": "SMS tanib bo'lmadi"}, status=200)

        matches = db.find_cards_by_last4(parsed.card_last4)
        if not matches:
            logger.info("SMSga mos karta topilmadi (oxirgi 4 raqam: %s)", parsed.card_last4)
            for admin_id in ADMIN_IDS:
                try:
                    await bot.send_message(
                        admin_id,
                        f"⚠️ Noma'lum kartaga SMS keldi (*{parsed.card_last4}):\n<code>{text[:300]}</code>",
                    )
                except Exception:
                    pass
            return web.json_response({"ok": False, "error": "Karta topilmadi"}, status=200)

        card = matches[0]
        # add_transaction avtomatik ravishda cards.balance ni ham yangilaydi
        db.add_transaction(card["id"], parsed.tx_type, parsed.amount, parsed.counterparty, text)
        final_balance = db.get_card(card["id"])["balance"]

        await _send_receipt(bot, card, parsed, final_balance)

        if parsed.tx_type == "expense" and card["owner_user_id"] is not None:
            await check_card_budget_and_warn(bot, card["id"], card["owner_user_id"])

        return web.json_response({"ok": True, "card_id": card["id"], "new_balance": final_balance})

    app.router.add_get("/", health)
    app.router.add_get("/health", health)
    app.router.add_post("/sms", sms_webhook)
    return app


async def _send_receipt(bot: Bot, card, parsed: ParsedSms, final_balance: float) -> None:
    """Kartaga bog'liq foydalanuvchiga va barcha adminlarga 'chek' ko'rinishidagi xabar yuboradi."""
    icon = "📥" if parsed.tx_type == "income" else "📤"
    action = "KIRIM" if parsed.tx_type == "income" else "CHIQIM"
    cp_line = f"👤 Kimga/kimdan: {parsed.counterparty}\n" if parsed.counterparty else ""

    receipt = (
        f"{icon} <b>{action} — CHEK</b>\n"
        f"━━━━━━━━━━━━━━━\n"
        f"💳 Karta: {card['card_type']} {mask_card_number(card['card_number'])}\n"
        f"💵 Summa: <b>{format_money(parsed.amount)}</b>\n"
        f"{cp_line}"
        f"💰 Qoldiq: <b>{format_money(final_balance)}</b>\n"
        f"━━━━━━━━━━━━━━━"
    )

    recipients = {*ADMIN_IDS}
    if card["owner_user_id"] is not None:
        recipients.add(card["owner_user_id"])
    for user_id in recipients:
        try:
            await bot.send_message(user_id, receipt)
        except Exception as e:
            logger.warning("Chek xabarini yuborib bo'lmadi (user_id=%s): %s", user_id, e)


# ======================================================================
# 8. HANDLERLAR
# ======================================================================
# --- handlers/common.py ---
router_common = Router(name="common")


@router_common.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    user = message.from_user
    db.upsert_user(user.id, user.username, user.full_name)

    text = (
        f"👋 Assalomu alaykum, <b>{escape_html(user.full_name)}</b>!\n\n"
        "💼 <b>Hamyonim</b> — kartalaringizni avtomatik kuzatib boradigan shaxsiy moliya botingiz.\n\n"
        "💳 Kartalaringiz admin tomonidan ulanadi, va har bir kirim/chiqim avtomatik qayd etiladi."
    )
    await message.answer(text, reply_markup=main_menu_keyboard(user.id))


@router_common.message(F.text == "🔄 Start")
async def btn_start(message: Message, state: FSMContext) -> None:
    await cmd_start(message, state)


@router_common.message(Command("cancel"))
@router_common.message(F.text == "🚫 Bekor qilish")
async def cmd_cancel(message: Message, state: FSMContext) -> None:
    current_state = await state.get_state()
    if current_state is None:
        await message.answer("❕ Hozircha bekor qilinadigan amal yo'q.", reply_markup=main_menu_keyboard(message.from_user.id))
        return
    await state.clear()
    await message.answer("✅ Amal bekor qilindi.", reply_markup=main_menu_keyboard(message.from_user.id))


@router_common.message(Command("balans"))
@router_common.message(F.text == "💰 Balans")
async def show_balance(message: Message) -> None:
    user_id = message.from_user.id
    if not db.user_exists(user_id):
        db.upsert_user(user_id, message.from_user.username, message.from_user.full_name)

    cards = db.get_cards_by_owner(user_id)
    total = db.get_total_balance(user_id)

    if not cards:
        await message.answer(
            "💳 Sizda hali karta yo'q.\n\n\"➕ Karta qo'shish\" tugmasini bosib, o'z kartangizni ulang.",
        )
        return

    emoji = "🟢" if total >= 0 else "🔴"
    lines = [f"{emoji} <b>Umumiy balansingiz:</b> {format_money(total)}\n"]
    for c in cards:
        lines.append(f"• {c['card_type']}: {format_money(c['balance'])}")

    await message.answer("\n".join(lines))


# --------------------------------------------------------- ADMIN BILAN BOG'LANISH

@router_common.message(F.text == "✉️ Admin bilan bog'lanish")
async def start_contact_admin(message: Message, state: FSMContext) -> None:
    if message.from_user.id in ADMIN_IDS:
        return  # Admin o'ziga yozmaydi
    await state.set_state(ContactAdminStates.entering_message)
    await message.answer(
        "✉️ Admin uchun xabaringizni yozing. U tez orada javob beradi.",
        reply_markup=cancel_keyboard(),
    )


@router_common.message(ContactAdminStates.entering_message)
async def relay_message_to_admin(message: Message, state: FSMContext, bot: Bot) -> None:
    await state.clear()
    user = message.from_user

    if not ADMIN_IDS:
        await message.answer("⚠️ Hozircha admin ulanmagan.", reply_markup=main_menu_keyboard(user.id))
        return

    header = f"✉️ <b>Yangi xabar</b>\n👤 {escape_html(user.full_name)} (@{user.username or '—'}, ID: <code>{user.id}</code>):\n"

    for admin_id in ADMIN_IDS:
        try:
            sent = await bot.send_message(admin_id, header + escape_html(message.text or ""))
            db.save_contact_relay(sent.message_id, admin_id, user.id)
        except Exception:
            continue

    await message.answer("✅ Xabaringiz adminga yuborildi!", reply_markup=main_menu_keyboard(user.id))


@router_common.message(F.reply_to_message, F.from_user.id.in_(ADMIN_IDS))
async def relay_admin_reply_to_user(message: Message, bot: Bot) -> None:
    """Admin forward qilingan xabarga reply qilsa, javob asl foydalanuvchiga yuboriladi."""
    source_user_id = db.get_contact_relay_source(message.reply_to_message.message_id, message.chat.id)
    if source_user_id is None:
        return  # Bu admin panelidagi oddiy reply, aloqaga tegishli emas

    try:
        await bot.send_message(source_user_id, f"✉️ <b>Admin javobi:</b>\n{message.text}")
        await message.answer("✅ Javobingiz foydalanuvchiga yuborildi.")
    except Exception as e:
        await message.answer(f"⚠️ Yuborib bo'lmadi: {e}")


# --------------------------------------------------------------- FALLBACK

@router_common.message(StateFilter(None))
async def fallback_handler(message: Message) -> None:
    if not db.user_exists(message.from_user.id):
        db.upsert_user(message.from_user.id, message.from_user.username, message.from_user.full_name)
    await message.answer(
        "❓ Kechirasiz, tushunmadim. Quyidagi menyudan foydalaning.",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


# --- handlers/cards.py ---
router_cards = Router(name="cards")


@router_cards.message(F.text.in_({"💳 Kartalarim", "🧾 Tarix"}))
async def show_cards(message: Message) -> None:
    user_id = message.from_user.id
    cards = db.get_cards_by_owner(user_id)

    if not cards:
        await message.answer(
            "💳 Sizda hali karta yo'q.\n\n"
            "\"➕ Karta qo'shish\" tugmasini bosib, o'z kartangizni ulang — "
            "shundan so'ng barcha kirim/chiqimlar shu yerda avtomatik ko'rinadi.",
            reply_markup=cards_list_keyboard(cards),
        )
        return

    text = "💳 <b>Sizning kartalaringiz:</b>\n\nKartani tanlang — uning kirim/chiqim tarixi ochiladi:"
    await message.answer(text, reply_markup=cards_list_keyboard(cards))


@router_cards.message(F.text == "➕ Karta qo'shish")
async def start_self_add_card(message: Message, state: FSMContext) -> None:
    if not db.user_exists(message.from_user.id):
        db.upsert_user(message.from_user.id, message.from_user.username, message.from_user.full_name)

    await state.set_state(SelfAddCardStates.entering_card_number)
    await message.answer(
        "💳 <b>Yangi karta qo'shish</b>\n\n"
        "Karta raqamingizni kiriting (16 ta raqam):\n"
        "<i>Masalan: 8600 1234 5678 9012</i>\n\n"
        "🔒 Karta raqamingiz xavfsiz saqlanadi.",
        reply_markup=cancel_keyboard(),
    )


@router_cards.callback_query(F.data == "self_add_card")
async def start_self_add_card_callback(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(SelfAddCardStates.entering_card_number)
    await callback.message.answer(
        "💳 <b>Yangi karta qo'shish</b>\n\n"
        "Karta raqamingizni kiriting (16 ta raqam):\n"
        "<i>Masalan: 8600 1234 5678 9012</i>\n\n"
        "🔒 Karta raqamingiz xavfsiz saqlanadi.",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@router_cards.message(SelfAddCardStates.entering_card_number)
async def process_self_card_number(message: Message, state: FSMContext) -> None:
    card_number = message.text.strip()
    user_id = message.from_user.id

    # Xavfsizlik: karta raqami yozilgan xabarni darhol o'chirishga harakat qilamiz
    try:
        await message.delete()
    except TelegramAPIError:
        pass

    if not is_valid_card_number(card_number):
        await message.answer(
            "⚠️ Noto'g'ri format. Karta raqami 16 ta raqamdan iborat bo'lishi kerak.\nQaytadan kiriting yoki bekor qiling:",
            reply_markup=cancel_keyboard(),
        )
        return

    await state.clear()

    # 1) Karta raqami bo'yicha DBdan qidiramiz — balki admin uni oldindan (egasiz) qo'shgan
    existing = db.find_card_by_full_number(card_number)

    if existing is not None:
        if existing["owner_user_id"] is None:
            # Karta hali hech kimga bog'lanmagan — aynan shu foydalanuvchiga avtomatik bog'laymiz
            db.assign_card_owner(existing["id"], user_id)
            await message.answer(
                f"✅ <b>Karta sizga bog'landi!</b>\n\n"
                f"💳 Turi: {existing['card_type']}\n"
                f"🔢 Raqami: {mask_card_number(existing['card_number'])}\n\n"
                f"Endi shu kartaning barcha kirim/chiqimlari sizga avtomatik ko'rinadi.",
                reply_markup=main_menu_keyboard(user_id),
            )
        elif existing["owner_user_id"] == user_id:
            await message.answer(
                "ℹ️ Bu karta allaqachon sizning ro'yxatingizda mavjud.",
                reply_markup=main_menu_keyboard(user_id),
            )
        else:
            # Xavfsizlik: boshqa foydalanuvchiga tegishli kartani o'zlashtirib olishga yo'l qo'ymaymiz
            await message.answer(
                "⛔ Bu karta allaqachon boshqa foydalanuvchiga bog'langan.\n"
                "Agar bu xatolik deb hisoblasangiz, \"✉️ Admin bilan bog'lanish\" orqali murojaat qiling.",
                reply_markup=main_menu_keyboard(user_id),
            )
        return

    # 2) Karta butunlay yangi — to'g'ridan-to'g'ri shu foydalanuvchiga qo'shamiz
    card_id = db.add_card(owner_user_id=user_id, card_number=card_number, added_by_admin=user_id)
    card = db.get_card(card_id)

    await message.answer(
        f"✅ <b>Karta muvaffaqiyatli qo'shildi!</b>\n\n"
        f"💳 Turi: {card['card_type']}\n"
        f"🔢 Raqami: {mask_card_number(card['card_number'])}\n\n"
        f"Endi shu kartaga tegishli bank SMS xabarlarini ulasangiz, "
        f"kirim/chiqimlar avtomatik qayd etiladi.",
        reply_markup=main_menu_keyboard(user_id),
    )


def _build_card_history_page(card_id: int, page: int) -> tuple[str, InlineKeyboardMarkup]:
    card = db.get_card(card_id)
    if card is None:
        return "⚠️ Karta topilmadi.", card_history_pagination_keyboard(card_id, 0, 1)

    total = db.count_transactions_by_card(card_id)
    total_pages = max((total + TRANSACTIONS_PER_PAGE - 1) // TRANSACTIONS_PER_PAGE, 1)
    page = max(0, min(page, total_pages - 1))

    rows = db.get_transactions_by_card(card_id, limit=TRANSACTIONS_PER_PAGE, offset=page * TRANSACTIONS_PER_PAGE)

    header = (
        f"{'💳' if card['card_type']=='Uzcard' else '🟣'} <b>{card['card_type']} {mask_card_number(card['card_number'])}</b>\n"
        f"💰 Joriy balans: <b>{format_money(card['balance'])}</b>\n\n"
    )

    if not rows:
        return header + "🧾 Bu kartada hali tranzaksiyalar yo'q.", card_history_pagination_keyboard(card_id, 0, 1)

    lines = [header + f"🧾 <b>Tarix</b> (sahifa {page + 1}/{total_pages}):\n"]
    for row in rows:
        icon = "📥" if row["type"] == "income" else "📤"
        sign = "+" if row["type"] == "income" else "-"
        cp = f" | {escape_html(row['counterparty'])}" if row["counterparty"] else ""
        lines.append(f"{icon} {sign}{format_money(row['amount'])}{cp}\n   🕒 {format_datetime(row['created_at'])}")

    return "\n".join(lines), card_history_pagination_keyboard(card_id, page, total_pages)


@router_cards.callback_query(F.data.startswith("card_hist:"))
async def show_card_history(callback: CallbackQuery) -> None:
    _, card_id_str, page_str = callback.data.split(":", maxsplit=2)
    card = db.get_card(int(card_id_str))

    if card is None or card["owner_user_id"] != callback.from_user.id:
        await callback.answer("⚠️ Bu karta sizga tegishli emas.", show_alert=True)
        return

    text, keyboard = _build_card_history_page(int(card_id_str), int(page_str))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard)
    except TelegramAPIError:
        pass
    await callback.answer()


@router_cards.callback_query(F.data == "cards_back")
async def cards_back(callback: CallbackQuery) -> None:
    cards = db.get_cards_by_owner(callback.from_user.id)
    text = "💳 <b>Sizning kartalaringiz:</b>\n\nKartani tanlang — uning kirim/chiqim tarixi ochiladi:"
    try:
        await callback.message.edit_text(text, reply_markup=cards_list_keyboard(cards))
    except TelegramAPIError:
        pass
    await callback.answer()


@router_cards.callback_query(F.data == "tx_noop")
async def noop(callback: CallbackQuery) -> None:
    await callback.answer()


# --- handlers/admin.py ---
router_admin = Router(name="admin")


def _is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


@router_admin.message(F.text == "🛠 Admin Panel")
async def show_admin_panel(message: Message) -> None:
    if not _is_admin(message.from_user.id):
        return
    await message.answer("🛠 <b>Admin Panel</b>\n\nKerakli bo'limni tanlang:", reply_markup=admin_panel_keyboard())


# --------------------------------------------------------------- STATISTIKA

@router_admin.callback_query(F.data == "admin_stats")
async def show_admin_stats(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return

    stats = db.get_global_stats()
    active_today = db.count_active_users_today()

    text = (
        "📊 <b>Umumiy statistika</b>\n\n"
        f"👥 Foydalanuvchilar: <b>{stats['total_users']}</b>\n"
        f"🟢 Bugun faol: <b>{active_today}</b>\n"
        f"🚫 Bloklangan: <b>{stats['blocked_users']}</b>\n\n"
        f"💳 Jami kartalar: <b>{stats['total_cards']}</b>\n"
        f"💰 Barcha kartalar balansi: <b>{format_money(stats['total_balance'])}</b>\n\n"
        f"🧾 Jami tranzaksiyalar: <b>{stats['total_transactions']}</b>\n"
        f"📥 Umumiy kirim: <b>{format_money(stats['total_income'])}</b>\n"
        f"📤 Umumiy chiqim: <b>{format_money(stats['total_expense'])}</b>"
    )
    try:
        await callback.message.edit_text(text, reply_markup=admin_panel_keyboard())
    except TelegramAPIError:
        await callback.message.answer(text, reply_markup=admin_panel_keyboard())
    await callback.answer()


@router_admin.callback_query(F.data == "admin_all_cards")
async def show_all_cards(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return

    cards = db.get_all_cards()
    if not cards:
        await callback.message.answer("💳 Hali birorta ham karta qo'shilmagan.", reply_markup=admin_panel_keyboard())
        await callback.answer()
        return

    lines = ["💳 <b>Barcha kartalar:</b>\n"]
    for c in cards:
        if c["owner_user_id"] is None:
            owner_name = "— (hali egasiz)"
        else:
            owner = db.get_user(c["owner_user_id"])
            owner_name = owner["full_name"] if owner else f"ID:{c['owner_user_id']}"
        lines.append(f"• {c['card_type']} {mask_card_number(c['card_number'])} — {owner_name} — {format_money(c['balance'])}")

    await callback.message.answer("\n".join(lines), reply_markup=admin_panel_keyboard())
    await callback.answer()


# ------------------------------------------------------------ KARTA QO'SHISH

@router_admin.callback_query(F.data == "admin_add_card")
async def start_add_card(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return

    await state.set_state(AdminAddCardStates.entering_card_number)
    await callback.message.answer(
        "💳 <b>Yangi karta qo'shish</b>\n\n"
        "Karta raqamini kiriting (16 ta raqam). Egasini ko'rsatish shart emas — "
        "haqiqiy egasi shu raqamni botga o'zi kiritganda, karta avtomatik unga bog'lanadi.",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@router_admin.message(AdminAddCardStates.entering_card_number)
async def process_card_number(message: Message, state: FSMContext) -> None:
    card_number = message.text.strip()

    # Xavfsizlik: karta raqami yozilgan xabarni darhol o'chiramiz
    try:
        await message.delete()
    except TelegramAPIError:
        pass

    if not is_valid_card_number(card_number):
        await message.answer("⚠️ Noto'g'ri format. Karta raqami 16 ta raqamdan iborat bo'lishi kerak.", reply_markup=cancel_keyboard())
        return

    await state.clear()

    existing = db.find_card_by_full_number(card_number)
    if existing is not None:
        await message.answer(
            f"⚠️ Bu karta allaqachon ro'yxatda bor (ID: {existing['id']}).",
            reply_markup=add_another_card_keyboard(),
        )
        return

    card_id = db.add_card(owner_user_id=None, card_number=card_number, added_by_admin=message.from_user.id)
    card = db.get_card(card_id)

    await message.answer(
        f"✅ <b>Karta ro'yxatga olindi (hali egasiz)!</b>\n\n"
        f"💳 Turi: {card['card_type']}\n"
        f"🔢 Raqami: {mask_card_number(card['card_number'])}\n\n"
        f"ℹ️ Karta egasi botga kirib, \"➕ Karta qo'shish\" orqali aynan shu raqamni "
        f"kiritishi bilanoq, bu karta avtomatik ravishda unga bog'lanadi.",
        reply_markup=add_another_card_keyboard(),
    )


@router_admin.callback_query(F.data == "admin_add_card_done")
async def finish_add_card(callback: CallbackQuery) -> None:
    await callback.message.answer("✅ Tugatildi.", reply_markup=main_menu_keyboard(callback.from_user.id))
    await callback.answer()


@router_admin.callback_query(F.data.startswith("admin_card_del:"))
async def delete_card(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return
    _, card_id_str = callback.data.split(":", maxsplit=1)
    success = db.delete_card(int(card_id_str))
    await callback.answer("✅ O'chirildi." if success else "⚠️ Topilmadi.")


# --------------------------------------------------------------- BROADCAST

@router_admin.callback_query(F.data == "admin_broadcast")
async def start_broadcast(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("⛔ Ruxsat yo'q.", show_alert=True)
        return
    await state.set_state(BroadcastStates.entering_message)
    await callback.message.answer("📢 Barcha foydalanuvchilarga yubormoqchi bo'lgan xabaringizni kiriting:", reply_markup=cancel_keyboard())
    await callback.answer()


@router_admin.message(BroadcastStates.entering_message)
async def process_broadcast_message(message: Message, state: FSMContext) -> None:
    await state.update_data(broadcast_text=message.html_text or message.text)
    await state.set_state(BroadcastStates.confirming)
    total_users = db.count_users()
    await message.answer(
        f"📋 <b>Xabar ko'rinishi:</b>\n\n{message.html_text or message.text}\n\n"
        f"👥 <b>{total_users}</b> ta foydalanuvchiga yuboriladi. Tasdiqlaysizmi?",
        reply_markup=broadcast_confirm_keyboard(),
    )


@router_admin.callback_query(BroadcastStates.confirming, F.data == "broadcast_confirm")
async def confirm_broadcast(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    data = await state.get_data()
    text = data.get("broadcast_text", "")
    await state.clear()
    await callback.message.edit_text("📤 Xabar tarqatish boshlandi...")
    await callback.answer()

    user_ids = db.get_all_user_ids(only_active=True)
    sent, failed = 0, 0

    for idx, uid in enumerate(user_ids, start=1):
        try:
            await bot.send_message(uid, text)
            sent += 1
        except TelegramForbiddenError:
            db.set_blocked(uid, True)
            failed += 1
        except TelegramRetryAfter as e:
            await asyncio.sleep(e.retry_after)
            try:
                await bot.send_message(uid, text)
                sent += 1
            except TelegramAPIError:
                failed += 1
        except TelegramAPIError as e:
            logger.warning("Broadcast xatoligi (user_id=%s): %s", uid, e)
            failed += 1

        if idx % USERS_PER_BROADCAST_BATCH == 0:
            await asyncio.sleep(BROADCAST_DELAY_SECONDS * USERS_PER_BROADCAST_BATCH)
        else:
            await asyncio.sleep(BROADCAST_DELAY_SECONDS)

    await callback.message.answer(f"✅ <b>Yakunlandi!</b>\n📤 Yuborildi: {sent}\n❌ Muvaffaqiyatsiz: {failed}")


@router_admin.callback_query(BroadcastStates.confirming, F.data == "broadcast_cancel")
async def cancel_broadcast(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await callback.message.edit_text("🚫 Bekor qilindi.")
    await callback.answer()


# --- handlers/budget.py ---
router_budget = Router(name="budget")


@router_budget.message(F.text == "🎯 Byudjet")
async def show_budget_menu(message: Message) -> None:
    user_id = message.from_user.id
    cards = db.get_cards_by_owner(user_id)

    if not cards:
        await message.answer("🎯 Byudjet belgilash uchun avval kartangizni qo'shing.")
        return

    items = []
    lines = ["🎯 <b>Byudjet nazorati</b>\n"]
    has_any = False
    for c in cards:
        budget = db.get_card_budget(c["id"])
        items.append((c, budget))
        if budget is not None:
            has_any = True
            spent = db.get_card_spent_this_month(c["id"])
            percent = (spent / budget["monthly_limit"] * 100) if budget["monthly_limit"] > 0 else 0
            icon = "🔴" if percent >= 100 else ("🟡" if percent >= 80 else "🟢")
            lines.append(
                f"{icon} {c['card_type']} {mask_card_number(c['card_number'])}: "
                f"{format_money(spent)} / {format_money(budget['monthly_limit'])} ({percent:.0f}%)"
            )

    if not has_any:
        lines.append("Hali byudjet belgilanmagan.")

    await message.answer("\n".join(lines), reply_markup=budget_list_keyboard(items))


@router_budget.callback_query(F.data == "budget_add")
async def start_add_budget(callback: CallbackQuery, state: FSMContext) -> None:
    cards = db.get_cards_by_owner(callback.from_user.id)
    if not cards:
        await callback.answer("Avval karta qo'shing.", show_alert=True)
        return
    await state.set_state(BudgetStates.choosing_card)
    await callback.message.answer("Qaysi karta uchun byudjet belgilaysiz?", reply_markup=budget_card_choice_keyboard(cards))
    await callback.answer()


@router_budget.callback_query(BudgetStates.choosing_card, F.data.startswith("budget_card:"))
async def choose_budget_card(callback: CallbackQuery, state: FSMContext) -> None:
    _, card_id_str = callback.data.split(":", maxsplit=1)
    await state.update_data(card_id=int(card_id_str))
    await state.set_state(BudgetStates.entering_limit)
    await callback.message.edit_text("💵 Oylik xarajat limitini kiriting (so'mda):")
    await callback.answer()


@router_budget.callback_query(F.data == "budget_cancel")
async def cancel_budget(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await callback.message.edit_text("🚫 Bekor qilindi.")
    await callback.answer()


@router_budget.message(BudgetStates.entering_limit)
async def process_budget_limit(message: Message, state: FSMContext) -> None:
    amount = safe_float(message.text)
    if amount is None or amount <= 0:
        await message.answer("⚠️ Musbat son kiriting, masalan: 500000")
        return

    data = await state.get_data()
    db.set_card_budget(data["card_id"], amount)
    await state.clear()

    await message.answer(
        f"✅ Byudjet saqlandi: {format_money(amount)}/oy",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


@router_budget.callback_query(F.data.startswith("budget_del:"))
async def delete_budget(callback: CallbackQuery) -> None:
    _, card_id_str = callback.data.split(":", maxsplit=1)
    db.delete_card_budget(int(card_id_str))
    await callback.answer("✅ Byudjet o'chirildi.")
    try:
        await callback.message.delete()
    except Exception:
        pass


async def check_card_budget_and_warn(bot, card_id: int, owner_user_id: int) -> None:
    """SMS orqali chiqim yozilgandan so'ng chaqiriladi — limitga yaqinlashilgan/oshib ketilganini tekshiradi."""
    budget = db.get_card_budget(card_id)
    if budget is None:
        return

    spent = db.get_card_spent_this_month(card_id)
    limit = budget["monthly_limit"]
    if limit <= 0:
        return

    ratio = spent / limit
    card = db.get_card(card_id)
    label = f"{card['card_type']} {mask_card_number(card['card_number'])}"

    if ratio >= BUDGET_EXCEEDED_THRESHOLD:
        text = f"🔴 <b>Byudjet oshib ketdi!</b>\n{label}: {format_money(spent)} / {format_money(limit)}"
    elif ratio >= BUDGET_WARNING_THRESHOLD:
        text = f"⚠️ <b>Byudjet ogohlantirishi!</b>\n{label}: {format_money(spent)} / {format_money(limit)} ({ratio*100:.0f}%)"
    else:
        return

    try:
        await bot.send_message(owner_user_id, text)
    except Exception:
        pass


# --- handlers/goals.py ---
router_goals = Router(name="goals")


def _progress_bar(current: float, target: float, length: int = 10) -> str:
    percent = min(current / target, 1.0) if target > 0 else 0.0
    filled = int(round(percent * length))
    return f"[{'█' * filled}{'░' * (length - filled)}] {percent * 100:.0f}%"


@router_goals.message(F.text == "🏆 Maqsadlar")
async def show_goals(message: Message) -> None:
    goals = db.get_goals(message.from_user.id)
    text = "🏆 <b>Jamg'arma maqsadlaringiz</b>" + ("" if goals else "\n\nHali maqsad yo'q. Yangi maqsad qo'shing!")
    await message.answer(text, reply_markup=goals_list_keyboard(goals))


@router_goals.callback_query(F.data == "goal_add")
async def start_add_goal(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(GoalStates.entering_name)
    await callback.message.answer("🏷 Maqsad nomini kiriting (masalan: Yangi noutbuk):", reply_markup=cancel_keyboard())
    await callback.answer()


@router_goals.message(GoalStates.entering_name)
async def process_goal_name(message: Message, state: FSMContext) -> None:
    await state.update_data(name=message.text.strip())
    await state.set_state(GoalStates.entering_target)
    await message.answer("💵 Maqsad summasini kiriting (so'mda):")


@router_goals.message(GoalStates.entering_target)
async def process_goal_target(message: Message, state: FSMContext) -> None:
    amount = safe_float(message.text)
    if amount is None or amount <= 0:
        await message.answer("⚠️ Musbat son kiriting.")
        return
    await state.update_data(target=amount)
    await state.set_state(GoalStates.entering_deadline)
    await message.answer("📅 Muddatni kiriting (KK.OO.YYYY) yoki '-' — muddatsiz:")


@router_goals.message(GoalStates.entering_deadline)
async def process_goal_deadline(message: Message, state: FSMContext) -> None:
    raw = message.text.strip()
    deadline = None
    if raw != "-":
        try:
            deadline = datetime.strptime(raw, "%d.%m.%Y").date().isoformat()
        except ValueError:
            await message.answer("⚠️ Format: KK.OO.YYYY (masalan 31.12.2026) yoki '-'")
            return

    data = await state.get_data()
    db.add_goal(message.from_user.id, data["name"], data["target"], deadline)
    await state.clear()
    await message.answer(
        f"✅ Yangi maqsad yaratildi!\n\n🏷 {data['name']}\n🎯 {format_money(data['target'])}",
        reply_markup=main_menu_keyboard(message.from_user.id),
    )


@router_goals.callback_query(F.data.startswith("goal_view:"))
async def view_goal(callback: CallbackQuery) -> None:
    _, goal_id_str = callback.data.split(":", maxsplit=1)
    goal = db.get_goal(int(goal_id_str), callback.from_user.id)
    if goal is None:
        await callback.answer("⚠️ Topilmadi.", show_alert=True)
        return

    bar = _progress_bar(goal["current_amount"], goal["target_amount"])
    text = (
        f"🏆 <b>{goal['name']}</b>\n\n{bar}\n"
        f"💰 {format_money(goal['current_amount'])} / {format_money(goal['target_amount'])}"
    )
    if goal["completed"]:
        text += "\n\n🎉 Tabriklaymiz, maqsadga erishdingiz!"

    try:
        await callback.message.edit_text(text, reply_markup=goal_detail_keyboard(goal["id"]))
    except Exception:
        await callback.message.answer(text, reply_markup=goal_detail_keyboard(goal["id"]))
    await callback.answer()


@router_goals.callback_query(F.data == "goal_back")
async def goal_back(callback: CallbackQuery) -> None:
    goals = db.get_goals(callback.from_user.id)
    try:
        await callback.message.edit_text("🏆 <b>Jamg'arma maqsadlaringiz</b>", reply_markup=goals_list_keyboard(goals))
    except Exception:
        pass
    await callback.answer()


@router_goals.callback_query(F.data.startswith("goal_deposit:"))
async def start_deposit(callback: CallbackQuery, state: FSMContext) -> None:
    _, goal_id_str = callback.data.split(":", maxsplit=1)
    await state.update_data(goal_id=int(goal_id_str))
    await state.set_state(GoalDepositStates.entering_amount)
    await callback.message.answer("💵 Qancha mablag' qo'shmoqchisiz?", reply_markup=cancel_keyboard())
    await callback.answer()


@router_goals.message(GoalDepositStates.entering_amount)
async def process_deposit(message: Message, state: FSMContext) -> None:
    amount = safe_float(message.text)
    if amount is None or amount <= 0:
        await message.answer("⚠️ Musbat son kiriting.")
        return

    data = await state.get_data()
    goal = db.add_to_goal(data["goal_id"], message.from_user.id, amount)
    await state.clear()

    if goal is None:
        await message.answer("⚠️ Maqsad topilmadi.", reply_markup=main_menu_keyboard(message.from_user.id))
        return

    bar = _progress_bar(goal["current_amount"], goal["target_amount"])
    await message.answer(f"✅ {format_money(amount)} qo'shildi!\n\n{bar}", reply_markup=main_menu_keyboard(message.from_user.id))
    if goal["completed"]:
        await message.answer(f"🎉 Tabriklaymiz! \"{goal['name']}\" maqsadingizga erishdingiz!")


@router_goals.callback_query(F.data.startswith("goal_del:"))
async def delete_goal(callback: CallbackQuery) -> None:
    _, goal_id_str = callback.data.split(":", maxsplit=1)
    db.delete_goal(int(goal_id_str), callback.from_user.id)
    goals = db.get_goals(callback.from_user.id)
    try:
        await callback.message.edit_text("🏆 <b>Jamg'arma maqsadlaringiz</b>", reply_markup=goals_list_keyboard(goals))
    except Exception:
        pass
    await callback.answer("✅ O'chirildi.")


# --- handlers/stats.py ---
import io
import re


router_stats = Router(name="stats")


@router_stats.message(F.text == "📊 Statistika")
async def show_stats_menu(message: Message) -> None:
    await message.answer("📊 <b>Statistika</b>\n\nQaysi davr uchun ko'rmoqchisiz?", reply_markup=stats_period_keyboard())


@router_stats.callback_query(F.data.startswith("stats_period:"))
async def show_stats(callback: CallbackQuery) -> None:
    _, days_str = callback.data.split(":", maxsplit=1)
    days = int(days_str)
    user_id = callback.from_user.id

    summary = db.get_stats_summary_for_user(user_id, days=days)
    expense_breakdown = db.get_counterparty_breakdown_for_user(user_id, "expense", days=days)
    income_breakdown = db.get_counterparty_breakdown_for_user(user_id, "income", days=days)

    net = summary["income"] - summary["expense"]
    net_emoji = "🟢" if net >= 0 else "🔴"

    lines = [
        f"📊 <b>Oxirgi {days} kunlik statistika</b>\n",
        f"📥 Jami kirim: <b>{format_money(summary['income'])}</b>",
        f"📤 Jami chiqim: <b>{format_money(summary['expense'])}</b>",
        f"{net_emoji} Sof natija: <b>{format_money(net)}</b>\n",
    ]

    if expense_breakdown:
        lines.append("📤 <b>Chiqimlar (kimga):</b>")
        for row in expense_breakdown[:6]:
            lines.append(f"  • {row['counterparty']}: {format_money(row['total'])} ({row['cnt']})")

    if income_breakdown:
        lines.append("\n📥 <b>Kirimlar (kimdan):</b>")
        for row in income_breakdown[:6]:
            lines.append(f"  • {row['counterparty']}: {format_money(row['total'])} ({row['cnt']})")

    text = "\n".join(lines)
    try:
        await callback.message.edit_text(text, reply_markup=stats_period_keyboard())
    except Exception:
        await callback.message.answer(text, reply_markup=stats_period_keyboard())
    await callback.answer()


@router_stats.callback_query(F.data == "stats_chart")
async def show_stats_chart(callback: CallbackQuery) -> None:
    await callback.answer("📊 Grafik tayyorlanmoqda...")
    user_id = callback.from_user.id

    breakdown = db.get_counterparty_breakdown_for_user(user_id, "expense", days=30)
    if not breakdown:
        await callback.message.answer("🧾 Oxirgi 30 kunda chiqimlar topilmadi.")
        return

    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    def clean_label(text: str) -> str:
        return re.sub(r"[^\w\s'\u0400-\u04FF-]", "", text).strip() or text

    labels = [clean_label(row["counterparty"]) for row in breakdown]
    values = [row["total"] for row in breakdown]

    fig, ax = plt.subplots(figsize=(6, 6))
    ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
    ax.axis("equal")
    ax.set_title("Chiqimlar taqsimoti (30 kun)")

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    buf.seek(0)

    photo = BufferedInputFile(buf.read(), filename="chiqimlar_grafigi.png")
    await callback.message.answer_photo(photo)


# --- handlers/export_data.py ---
import io


router_export = Router(name="export_data")


@router_export.message(F.text == "📁 Eksport")
async def show_export_menu(message: Message) -> None:
    await message.answer("📁 <b>Ma'lumotlarni eksport qilish</b>\n\nQaysi formatda yuklab olasiz?", reply_markup=export_format_keyboard())


@router_export.callback_query(F.data == "export_xlsx")
async def export_excel(callback: CallbackQuery) -> None:
    await callback.answer("⏳ Tayyorlanmoqda...")
    user_id = callback.from_user.id
    rows = db.get_all_transactions_for_user(user_id)

    if not rows:
        await callback.message.answer("📁 Eksport qilish uchun ma'lumot topilmadi.")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Tranzaksiyalar"
    ws.append(["ID", "Sana", "Turi", "Summa", "Karta", "Kimga/kimdan"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

    for row in rows:
        ws.append([
            row["id"], format_datetime(row["created_at"]),
            "Kirim" if row["type"] == "income" else "Chiqim",
            row["amount"], f"{row['card_type']} {mask_card_number(row['card_number'])}",
            row["counterparty"] or "",
        ])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename="hamyonim_tranzaksiyalar.xlsx"),
        caption="📁 Hisobotingiz tayyor!",
    )


@router_export.callback_query(F.data == "export_pdf")
async def export_pdf(callback: CallbackQuery) -> None:
    await callback.answer("⏳ Tayyorlanmoqda...")
    user_id = callback.from_user.id
    rows = db.get_all_transactions_for_user(user_id)

    if not rows:
        await callback.message.answer("📁 Eksport qilish uchun ma'lumot topilmadi.")
        return

    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Hamyonim - Tranzaksiyalar hisoboti", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", size=9)
    pdf.ln(2)

    col_widths = [12, 32, 20, 30, 45, 45]
    headers = ["ID", "Sana", "Turi", "Summa", "Karta", "Kimga/kimdan"]
    pdf.set_fill_color(46, 125, 50)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 8, h, border=1, fill=True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    for row in rows:
        values = [
            str(row["id"]), format_datetime(row["created_at"]),
            "Kirim" if row["type"] == "income" else "Chiqim",
            format_money(row["amount"]),
            _safe(f"{row['card_type']} {mask_card_number(row['card_number'])}"),
            _safe((row["counterparty"] or "")[:28]),
        ]
        for w, v in zip(col_widths, values):
            pdf.cell(w, 7, v, border=1)
        pdf.ln()

    pdf_bytes = bytes(pdf.output())
    await callback.message.answer_document(
        BufferedInputFile(pdf_bytes, filename="hamyonim_hisobot.pdf"),
        caption="📁 Hisobotingiz tayyor!",
    )


def _safe(text: str) -> str:
    try:
        text.encode("latin-1")
        return text
    except UnicodeEncodeError:
        return text.encode("latin-1", errors="replace").decode("latin-1")


# ======================================================================
# 9. ASOSIY ISHGA TUSHIRISH
# ======================================================================
async def set_bot_commands(bot: Bot) -> None:
    commands = [
        BotCommand(command="start", description="Botni ishga tushirish"),
        BotCommand(command="balans", description="Balansni ko'rish"),
        BotCommand(command="cancel", description="Bekor qilish"),
    ]
    await bot.set_my_commands(commands)


async def run_web_server(bot: Bot) -> None:
    """Render (yoki boshqa hosting) uchun portni ochib turadigan yengil server."""
    app = create_web_app(bot)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, host="0.0.0.0", port=PORT)
    await site.start()
    logger.info("Veb-server %s portida ishga tushdi (health-check + /sms webhook).", PORT)


async def main() -> None:
    if not BOT_TOKEN or ":" not in BOT_TOKEN:
        logger.error("BOT_TOKEN noto'g'ri sozlangan!")
        sys.exit(1)

    if not ADMIN_IDS:
        logger.warning(
            "ADMIN_IDS bo'sh! HAMYONIM_ADMIN_IDS muhit o'zgaruvchisiga o'z Telegram ID'ingizni yozing, "
            "aks holda Admin Panelga hech kim kira olmaydi."
        )

    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp = Dispatcher(storage=MemoryStorage())

    for r in (router_admin, router_cards, router_budget, router_goals, router_stats, router_export, router_common):
        dp.include_router(r)

    await set_bot_commands(bot)

    logger.info("=" * 60)
    logger.info("HAMYONIM BOTI ISHGA TUSHMOQDA...")
    logger.info("Adminlar: %s", ADMIN_IDS or "(belgilanmagan)")
    logger.info("SMS webhook manzili: http://<sizning-hostingiz>/sms")
    logger.info("SMS webhook maxfiy kaliti: %s", SMS_WEBHOOK_SECRET)
    logger.info("=" * 60)

    await run_web_server(bot)

    try:
        await bot.delete_webhook(drop_pending_updates=True)
        await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())
    finally:
        await bot.session.close()
        logger.info("Bot to'xtatildi.")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Bot foydalanuvchi tomonidan to'xtatildi.")
